from bs4 import BeautifulSoup as bs
from requests.sessions import Session
from typing import Tuple, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, numbers
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

import os, datetime, smtplib, getpass

headers = {
    "accept": "*/*",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36",
}


def indicative_exchange_rate_parser(base_url: str, headers: dict) -> Tuple[List[str], List[float]]:
    session = Session()
    request = session.get(base_url, headers=headers)

    if request.status_code == 200:
        soup = bs(request.content, "html.parser")
        rates = soup.find_all("rate")
        date_x_value = []
        for rate in rates:
            date_x_value.append([rate["moment"], float(rate["value"])])

    date_x_value = date_x_value[::-1]
    return [date[0] for date in date_x_value], [value[1] for value in date_x_value]


today = datetime.date.today().strftime("%Y-%m-%d")
one_month_ago = (
    datetime.date.today().replace(month=datetime.date.today().month - 1).strftime("%Y-%m-%d")
)
base_url = "https://www.moex.com/export/derivatives/currency-rate.aspx?language=ru&currency={}&moment_start={}&moment_end={}"
dates, values = indicative_exchange_rate_parser(
    base_url.format("USD_RUB", one_month_ago, today),
    headers,
)  # USD/RUB

diff_values = ["-"]

for i in range(1, len(values)):
    diff_values.append(values[i] - values[i - 1])

excel_file = Workbook()
work_sheet = excel_file.active
work_sheet.cell(row=1, column=1, value="Дата").alignment = Alignment(
    horizontal="center", vertical="center"
)


for col_index, value in enumerate(dates):
    cell = work_sheet.cell(row=col_index + 2, column=1, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = numbers.FORMAT_DATE_DATETIME

work_sheet.cell(row=1, column=2, value="Курс").alignment = Alignment(
    horizontal="center", vertical="center"
)

for col_index, value in enumerate(values):
    cell = work_sheet.cell(row=col_index + 2, column=2, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '#,##0.00_-"₽"'

work_sheet.cell(row=1, column=3, value="Изменение").alignment = Alignment(
    horizontal="center", vertical="center"
)

for col_index, value in enumerate(diff_values):
    cell = work_sheet.cell(row=col_index + 2, column=3, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '#,##0.00_-"₽"'

dates_eur, values_eur = indicative_exchange_rate_parser(
    base_url.format("EUR_RUB", one_month_ago, today),
    headers,
)  # EUR/RUB

diff_values_eur = ["-"]

for i in range(1, len(values_eur)):
    diff_values_eur.append(values_eur[i] - values_eur[i - 1])

work_sheet.cell(row=1, column=4, value="Дата").alignment = Alignment(
    horizontal="center", vertical="center"
)

for col_index, value in enumerate(dates_eur):
    cell = work_sheet.cell(row=col_index + 2, column=4, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = numbers.FORMAT_DATE_DATETIME

work_sheet.cell(row=1, column=5, value="Курс").alignment = Alignment(
    horizontal="center", vertical="center"
)

for col_index, value in enumerate(values_eur):
    cell = work_sheet.cell(row=col_index + 2, column=5, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '#,##0.00_-"₽"'

work_sheet.cell(row=1, column=6, value="Изменение").alignment = Alignment(
    horizontal="center", vertical="center"
)

for col_index, value in enumerate(diff_values_eur):
    cell = work_sheet.cell(row=col_index + 2, column=6, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '#,##0.00_-"₽"'

EUR_to_USD = [eur / usd for eur, usd in zip(values_eur, values)]

work_sheet.cell(row=1, column=7, value="EUR/USD").alignment = Alignment(
    horizontal="center", vertical="center"
)

for col_index, value in enumerate(EUR_to_USD):
    cell = work_sheet.cell(row=col_index + 2, column=7, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '#,##0.00_-"₽"'

for column_cells in work_sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells) + 1
    work_sheet.column_dimensions[column_cells[0].column_letter].width = length

excel_file.save(os.path.join(os.getcwd(), "data.xlsx"))

email = "aaron.kuzhelev@gmail.com"

postfix = ""

if (len(values) + 1) % 10 == 1:
    postfix = "а"
elif (len(values) + 1) < 5 and (len(values) + 1) != 0:
    postfix = "и"
else:
    postfix = ""


num_of_rows = str((len(values) + 1)) + " cтрок" + postfix

msg = MIMEMultipart()
msg["From"] = email
msg["To"] = email
msg["Date"] = formatdate(localtime=True)
msg["Subject"] = "Отчет"
msg.attach(MIMEText(num_of_rows))

part = MIMEBase("application", "octet-stream")
part.set_payload(open(os.path.join(os.getcwd(), "data.xlsx"), "rb").read())
encoders.encode_base64(part)
part.add_header("Content-Disposition", 'attachment; filename="rate.xlsx"')
msg.attach(part)

smtp = smtplib.SMTP("smtp.gmail.com", 587)
smtp.starttls()
smtp.login(email, getpass.getpass("Input password: "))
smtp.sendmail(email, email, msg.as_string())
smtp.quit()